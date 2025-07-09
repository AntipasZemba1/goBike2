# import tkinter as tk
# from tkinter import ttk, messagebox
# from openpyxl import Workbook, load_workbook
# import os
# from datetime import datetime

# # Excel file to store the data
# DATA_FILE = "motor_taxi_data.xlsx"

# # Create workbook and sheet with headers if file doesn't exist
# if not os.path.exists(DATA_FILE):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "MotorTaxiData"
#     ws.append(["Date", "Driver Name", "Taxi ID", "Daily Earnings ($)", "Maintenance Notes"])
#     wb.save(DATA_FILE)

# # Function to save data to Excel
# def save_data():
#     date = date_entry.get()
#     driver = driver_entry.get()
#     taxi_id = taxi_id_entry.get()
#     earnings = earnings_entry.get()
#     notes = notes_entry.get()

#     if not (date and driver and taxi_id and earnings):
#         messagebox.showerror("Error", "Please fill in all required fields.")
#         return

#     try:
#         earnings_float = float(earnings)
#     except ValueError:
#         messagebox.showerror("Error", "Daily earnings must be a number.")
#         return

#     try:
#         wb = load_workbook(DATA_FILE)
#         ws = wb.active
#         ws.append([date, driver, taxi_id, earnings_float, notes])
#         wb.save(DATA_FILE)
#         messagebox.showinfo("Success", "Data saved successfully to Excel!")
#         clear_fields()
#     except Exception as e:
#         messagebox.showerror("Error", f"Failed to save data: {e}")

# # Function to clear fields after saving
# def clear_fields():
#     date_entry.delete(0, tk.END)
#     driver_entry.delete(0, tk.END)
#     taxi_id_entry.delete(0, tk.END)
#     earnings_entry.delete(0, tk.END)
#     notes_entry.delete(0, tk.END)

# # Set up the main Tkinter window
# root = tk.Tk()
# root.title("Motor Taxi Business Data Entry")
# root.geometry("400x400")
# root.resizable(False, False)

# # Title Label
# ttk.Label(root, text="Motor Taxi Data Entry", font=("Arial", 16)).pack(pady=10)

# # Form Fields
# ttk.Label(root, text="Date (YYYY-MM-DD):").pack(pady=5)
# date_entry = ttk.Entry(root)
# date_entry.insert(0, datetime.today().strftime('%Y-%m-%d'))
# date_entry.pack()

# ttk.Label(root, text="Driver Name:").pack(pady=5)
# driver_entry = ttk.Entry(root)
# driver_entry.pack()

# ttk.Label(root, text="Taxi ID:").pack(pady=5)
# taxi_id_entry = ttk.Entry(root)
# taxi_id_entry.pack()

# ttk.Label(root, text="Daily Earnings ($):").pack(pady=5)
# earnings_entry = ttk.Entry(root)
# earnings_entry.pack()

# ttk.Label(root, text="Maintenance Notes:").pack(pady=5)
# notes_entry = ttk.Entry(root)
# notes_entry.pack()

# # Save Button
# save_button = ttk.Button(root, text="Save Data", command=save_data)
# save_button.pack(pady=20)

# # Start the application
# root.mainloop()

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from fpdf import FPDF  # fpdf2 works with the same import
import os
from datetime import datetime
from collections import defaultdict

# Excel file to store the data
DATA_FILE = "motor_taxi_data.xlsx"
PASSWORD = "admin123"  # Change this to your preferred password

# Preload driver names
DRIVER_LIST = [
    "Driver 1", "Driver 2", "Driver 3", "Driver 4", "Driver 5",
    "Driver 6", "Driver 7", "Driver 8", "Driver 9", "Driver 10",
    "Driver 11", "Driver 12", "Driver 13", "Driver 14", "Driver 15",
    "Driver 16", "Driver 17", "Driver 18", "Driver 19", "Driver 20"
]

# Create workbook and sheet with headers if file doesn't exist
if not os.path.exists(DATA_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "MotorTaxiData"
    ws.append(["Date", "Driver Name", "Taxi ID", "Daily Earnings ($)", "Maintenance Notes"])
    wb.save(DATA_FILE)

# Function to save data to Excel
def save_data():
    date = date_entry.get()
    driver = driver_combo.get()
    taxi_id = taxi_id_entry.get()
    earnings = earnings_entry.get()
    notes = notes_entry.get()

    if not (date and driver and taxi_id and earnings):
        messagebox.showerror("Error", "Please fill in all required fields.")
        return

    try:
        earnings_float = float(earnings)
    except ValueError:
        messagebox.showerror("Error", "Daily earnings must be a number.")
        return

    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        ws.append([date, driver, taxi_id, earnings_float, notes])
        wb.save(DATA_FILE)
        messagebox.showinfo("Success", "Data saved successfully to Excel!")
        clear_fields()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")

# Clear form fields
def clear_fields():
    date_entry.delete(0, tk.END)
    driver_combo.set('')
    taxi_id_entry.delete(0, tk.END)
    earnings_entry.delete(0, tk.END)
    notes_entry.delete(0, tk.END)

# Display saved data in a popup window
def view_data():
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        data_window = tk.Toplevel(root)
        data_window.title("Saved Data")
        tree = ttk.Treeview(data_window, columns=("Date", "Driver", "Taxi ID", "Earnings", "Notes"), show='headings')
        tree.pack(expand=True, fill='both')

        # Define headings
        for col in tree["columns"]:
            tree.heading(col, text=col)

        # Insert data
        for row in ws.iter_rows(min_row=2, values_only=True):
            tree.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load data: {e}")

# Calculate weekly/monthly summaries and save to Excel
def generate_summary():
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        summary = defaultdict(float)

        for row in ws.iter_rows(min_row=2, values_only=True):
            date_str, _, _, earnings, _ = row
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            week_key = date_obj.strftime("%Y-W%U")
            month_key = date_obj.strftime("%Y-%m")

            summary[week_key] += earnings
            summary[month_key] += earnings

        summary_ws = wb.create_sheet(title="Summary")
        summary_ws.append(["Period", "Total Earnings ($)"])
        bold_font = Font(bold=True)
        summary_ws["A1"].font = bold_font
        summary_ws["B1"].font = bold_font

        for period, total in summary.items():
            summary_ws.append([period, total])

        wb.save(DATA_FILE)
        messagebox.showinfo("Success", "Summary generated and saved to Excel!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate summary: {e}")

# Export data to PDF
def export_to_pdf():
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Motor Taxi Data Report", ln=True, align='C')
        pdf.ln(10)

        # Add table header
        headers = ["Date", "Driver Name", "Taxi ID", "Daily Earnings ($)", "Maintenance Notes"]
        for header in headers:
            pdf.cell(40, 10, txt=header, border=1)
        pdf.ln()

        # Add data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            for item in row:
                pdf.cell(40, 10, txt=str(item), border=1)
            pdf.ln()

        pdf.output("motor_taxi_report.pdf")
        messagebox.showinfo("Success", "Data exported to PDF as motor_taxi_report.pdf!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export PDF: {e}")

# Password prompt
def check_password():
    password = simpledialog.askstring("Password", "Enter admin password:", show="*")
    if password != PASSWORD:
        messagebox.showerror("Access Denied", "Incorrect password. Exiting application.")
        root.destroy()

# Setup Tkinter window
root = tk.Tk()
root.title("Motor Taxi Business Management")
root.geometry("450x500")
root.resizable(False, False)

# Title Label
ttk.Label(root, text="Motor Taxi Data Entry", font=("Arial", 16)).pack(pady=10)

# Form Fields
ttk.Label(root, text="Date (YYYY-MM-DD):").pack(pady=5)
date_entry = ttk.Entry(root)
date_entry.insert(0, datetime.today().strftime('%Y-%m-%d'))
date_entry.pack()

ttk.Label(root, text="Driver Name:").pack(pady=5)
driver_combo = ttk.Combobox(root, values=DRIVER_LIST, state="readonly")
driver_combo.pack()

ttk.Label(root, text="Taxi ID:").pack(pady=5)
taxi_id_entry = ttk.Entry(root)
taxi_id_entry.pack()

ttk.Label(root, text="Daily Earnings ($):").pack(pady=5)
earnings_entry = ttk.Entry(root)
earnings_entry.pack()

ttk.Label(root, text="Maintenance Notes:").pack(pady=5)
notes_entry = ttk.Entry(root)
notes_entry.pack()

# Buttons
ttk.Button(root, text="Save Data", command=save_data).pack(pady=10)
ttk.Button(root, text="View Data", command=view_data).pack(pady=5)
ttk.Button(root, text="Generate Summary", command=generate_summary).pack(pady=5)
ttk.Button(root, text="Export to PDF", command=export_to_pdf).pack(pady=5)

# Run password check
root.after(100, check_password)
root.mainloop()
